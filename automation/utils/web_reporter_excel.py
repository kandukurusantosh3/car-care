# automation/utils/web_reporter_excel.py
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_web_excel_reports(results, output_dir):
    """
    Generates all requested Excel files and summaries for Selenium Web tests:
    - Automation_Test_Report.xlsx (All sheets)
    - Passed_Test_Cases.xlsx
    - Failed_Test_Cases.xlsx
    - Summary_Report.xlsx
    """
    os.makedirs(output_dir, exist_ok=True)
    
    wb = openpyxl.Workbook()
    
    # Fonts & Styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Calibri", size=11)
    title_font = Font(name="Calibri", size=14, bold=True, color="004B49") # Dark Teal
    
    header_fill = PatternFill(start_color="004B49", end_color="004B49", fill_type="solid")
    pass_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    fail_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    skip_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    # Sheet 1: Executed Test Cases
    ws_all = wb.active
    ws_all.title = "Executed Test Cases"
    
    # Sheet 2: Passed Tests
    ws_pass = wb.create_sheet(title="Passed Tests")
    
    # Sheet 3: Failed Tests
    ws_fail = wb.create_sheet(title="Failed Tests")
    
    # Sheet 4: Skipped Tests
    ws_skip = wb.create_sheet(title="Skipped Tests")

    headers = ["Test ID", "Module", "Test Name", "Status", "Execution Time (ms)", "Priority", "Failure Reason"]
    
    for ws in [ws_all, ws_pass, ws_fail, ws_skip]:
        ws.append(headers)
        ws.row_dimensions[1].height = 24
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center

    passed_count = 0
    failed_count = 0
    skipped_count = 0
    defect_list = []

    for r in results:
        status = r["status"]
        row_data = [
            r["test_id"],
            r["module"],
            r["name"],
            status,
            r.get("duration_ms", 35),
            r["priority"],
            r.get("error", "")
        ]
        
        # Append to Executed sheet
        ws_all.append(row_data)
        
        # Append to split sheets
        if status == "PASS":
            ws_pass.append(row_data)
            passed_count += 1
        elif status == "FAIL":
            ws_fail.append(row_data)
            failed_count += 1
            defect_list.append({"id": r["test_id"], "module": r["module"], "error": r["error"]})
        else:
            ws_skip.append(row_data)
            skipped_count += 1

    # Format data rows and apply color fills based on status
    for ws in [ws_all, ws_pass, ws_fail, ws_skip]:
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
            status_cell = ws.cell(row=row, column=4)
            status_val = status_cell.value
            
            fill = pass_fill if status_val == "PASS" else fail_fill if status_val == "FAIL" else skip_fill
            
            for col in range(1, len(headers) + 1):
                c = ws.cell(row=row, column=col)
                c.font = cell_font
                c.border = thin_border
                if col in [1, 4, 6]:
                    c.alignment = align_center
                else:
                    c.alignment = align_left
            status_cell.fill = fill

    total_executed = passed_count + failed_count + skipped_count

    # Sheet 5: Execution Metrics
    ws_metrics = wb.create_sheet(title="Execution Metrics")
    ws_metrics.append(["Execution Metrics Summary"])
    ws_metrics.cell(row=1, column=1).font = title_font
    ws_metrics.append([])
    ws_metrics.append(["Metric Name", "Value"])
    ws_metrics.append(["Total Executed Cases", total_executed])
    ws_metrics.append(["Passed Count", passed_count])
    ws_metrics.append(["Failed Count", failed_count])
    ws_metrics.append(["Skipped Count", skipped_count])
    ws_metrics.append(["Pass Rate (%)", f"{round((passed_count / max(1, total_executed)) * 100, 2)}%"])
    
    for row in range(3, 9):
        ws_metrics.row_dimensions[row].height = 20
        for col in range(1, 3):
            c = ws_metrics.cell(row=row, column=col)
            c.font = cell_font
            c.border = thin_border
            if row == 3:
                c.font = Font(name="Calibri", size=11, bold=True)
                c.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Sheet 6: Defect Summary
    ws_defects = wb.create_sheet(title="Defect Summary")
    ws_defects.append(["Defect ID", "Module", "Description / Failure Traceback"])
    ws_defects.row_dimensions[1].height = 24
    for col in range(1, 4):
        c = ws_defects.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = align_center

    for idx, defect in enumerate(defect_list, 2):
        ws_defects.append([defect["id"], defect["module"], defect["error"]])
        ws_defects.row_dimensions[idx].height = 20
        for col in range(1, 4):
            c = ws_defects.cell(row=idx, column=col)
            c.font = cell_font
            c.border = thin_border
            c.alignment = align_left

    # Set column widths dynamically for all sheets
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    master_path = os.path.join(output_dir, "Automation_Test_Report.xlsx")
    wb.save(master_path)

    # 2. Generate Passed_Test_Cases.xlsx
    wb_pass = openpyxl.Workbook()
    ws_p = wb_pass.active
    ws_p.title = "Passed"
    ws_p.append(headers)
    for r in results:
        if r["status"] == "PASS":
            ws_p.append([r["test_id"], r["module"], r["name"], r["status"], r.get("duration_ms", 35), r["priority"], r.get("error", "")])
    wb_pass.save(os.path.join(output_dir, "Passed_Test_Cases.xlsx"))

    # 3. Generate Failed_Test_Cases.xlsx
    wb_fail = openpyxl.Workbook()
    ws_f = wb_fail.active
    ws_f.title = "Failed"
    ws_f.append(headers)
    for r in results:
        if r["status"] == "FAIL":
            ws_f.append([r["test_id"], r["module"], r["name"], r["status"], r.get("duration_ms", 35), r["priority"], r.get("error", "")])
    wb_fail.save(os.path.join(output_dir, "Failed_Test_Cases.xlsx"))

    # 4. Generate Summary_Report.xlsx
    wb_sum = openpyxl.Workbook()
    ws_s = wb_sum.active
    ws_s.title = "Summary"
    ws_s.append(["Metric Name", "Value"])
    ws_s.append(["Total Executed", total_executed])
    ws_s.append(["Total Passed", passed_count])
    ws_s.append(["Total Failed", failed_count])
    ws_s.append(["Total Skipped", skipped_count])
    ws_s.append(["Success Rate", f"{round((passed_count / max(1, total_executed)) * 100, 2)}%"])
    wb_sum.save(os.path.join(output_dir, "Summary_Report.xlsx"))

    print(f"[REPORTER] Web Excel reports successfully generated in {output_dir}")
