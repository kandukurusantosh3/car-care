import openpyxl
from openpyxl.styles import Font, PatternFill
import os
from datetime import datetime

def generate_excel_report(test_results, output_dir="."):
    """
    Generates an Excel report for the Appium end-to-end tests.
    
    :param test_results: List of dictionaries with keys: 'step', 'status', 'error'
    :param output_dir: Directory where the report will be saved.
    """
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Execution Report"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    pass_font = Font(color="006100")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    fail_font = Font(color="9C0006")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Add Headers
    headers = ["Test Step", "Status", "Error Details"]
    ws.append(headers)
    
    for col in range(1, 4):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill

    # Add Data
    for result in test_results:
        step = result.get('step', 'Unknown Step')
        status = result.get('status', 'FAIL')
        error = result.get('error', '')
        
        ws.append([step, status, error])
        
        # Color coding
        current_row = ws.max_row
        status_cell = ws.cell(row=current_row, column=2)
        
        if status == 'PASS':
            status_cell.font = pass_font
            status_cell.fill = pass_fill
        else:
            status_cell.font = fail_font
            status_cell.fill = fail_fill

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Ensure output directory exists
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Save file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"Appium_Test_Report_{timestamp}.xlsx"
    file_path = os.path.join(output_dir, file_name)
    
    wb.save(file_path)
    print(f"Excel report generated successfully at: {file_path}")
    return file_path
